# using openpyxl library
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    # change it to reference the param to function
    workbook = xl.load_workbook('transactions.xlsx')
    sheet = workbook['Sheet1']
    # cell = sheet['a1'] # Access with name of the cell
    # cell = sheet.cell(1, 1) # Access with cell method on sheet
    # cell.value to access the value at the cell position

    # Loop through the sheet to access each row values
    for row in range(2, sheet.max_row + 1):  # Ignore the first row which is a heading
        cell = sheet.cell(row, 3)  # Access third column in the row
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(
        sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()

    chart.add_data(values)

    sheet.add_chart(chart, 'a7')

    # Can be overridden or saved to new file
    workbook.save('transactions-fixed.xlsx')
